"""JobTrackerTalent -- track job applications via voice commands.

Stores applications and follow-ups in a dedicated SQLite database
(default: data/job_tracker.db). Supports full CRUD, status tracking,
follow-up reminders, statistics, and XLSX export for unemployment
reporting.

Integration: Cowork can call add_from_cowork() and
get_active_applications() via the cowork_bridge.

Examples:
    "add a job application at Netflix for VP of Engineering"
    "I applied to the Microsoft Azure CISO role"
    "update the Netflix application to interviewing"
    "show my active applications"
    "what jobs need follow up"
    "how many applications this week"
    "mark the Google role as rejected"
    "show all jobs from LinkedIn"
    "export my job tracker"
    "find applications for engineer"
    "add a follow up for Netflix in 5 days"
"""
from __future__ import annotations

import json
import os
import re
import sqlite3
from datetime import datetime, date, timedelta
from typing import Any

from talents.base import BaseTalent
from core.llm_client import LLMError

import logging
log = logging.getLogger(__name__)

# ── Valid statuses and transitions ────────────────────────────────────────────

VALID_STATUSES = ("new", "applied", "interviewing", "offered", "rejected", "withdrawn")

# Each status maps to the set of statuses it can transition TO.
_STATUS_TRANSITIONS: dict[str, set[str]] = {
    "new":          {"applied", "withdrawn"},
    "applied":      {"interviewing", "offered", "rejected", "withdrawn"},
    "interviewing": {"offered", "rejected", "withdrawn"},
    "offered":      {"rejected", "withdrawn"},
    "rejected":     set(),
    "withdrawn":    set(),
}


def _data_dir() -> str:
    """Ensure data/ directory exists and return its path."""
    d = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
    os.makedirs(d, exist_ok=True)
    return d


def _normalize_company(name: str) -> str:
    """Normalize a company name for fuzzy matching."""
    n = name.lower().strip()
    # Strip common suffixes
    for suffix in (" inc", " inc.", " llc", " corp", " corp.",
                   " co", " co.", " ltd", " ltd.", " limited",
                   " corporation", " incorporated"):
        if n.endswith(suffix):
            n = n[: -len(suffix)].rstrip(",. ")
    # Strip leading "the"
    if n.startswith("the "):
        n = n[4:]
    return n.strip()


def _friendly_date(iso_str: str | None) -> str:
    """Format an ISO date string as a human-friendly string."""
    if not iso_str:
        return ""
    try:
        d = date.fromisoformat(iso_str[:10])
    except (ValueError, TypeError):
        return iso_str
    today = date.today()
    delta = (today - d).days
    if delta == 0:
        return "today"
    elif delta == 1:
        return "yesterday"
    elif delta < 7:
        return f"{delta} days ago"
    else:
        return d.strftime("%b %d")


def _today_iso() -> str:
    return date.today().isoformat()


# ── Database helpers ──────────────────────────────────────────────────────────

_SCHEMA_APPLICATIONS = """\
CREATE TABLE IF NOT EXISTS applications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company TEXT NOT NULL,
    position TEXT NOT NULL,
    location TEXT DEFAULT '',
    source TEXT DEFAULT '',
    status TEXT DEFAULT 'new',
    date_found TEXT,
    date_applied TEXT,
    date_updated TEXT,
    contact_name TEXT DEFAULT '',
    contact_email TEXT DEFAULT '',
    method TEXT DEFAULT '',
    salary_range TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    job_url TEXT DEFAULT '',
    resume_version TEXT DEFAULT '',
    cover_letter INTEGER DEFAULT 0,
    cowork_task_id TEXT DEFAULT '',
    fit_score INTEGER DEFAULT 0,
    archived INTEGER DEFAULT 0
)"""

_SCHEMA_FOLLOW_UPS = """\
CREATE TABLE IF NOT EXISTS follow_ups (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    application_id INTEGER NOT NULL,
    due_date TEXT NOT NULL,
    note TEXT DEFAULT '',
    completed INTEGER DEFAULT 0,
    FOREIGN KEY (application_id) REFERENCES applications(id)
)"""


class _DB:
    """Thin SQLite wrapper scoped to a single database file."""

    def __init__(self, db_path: str) -> None:
        self.db_path = db_path
        os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)
        self._init_schema()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def _init_schema(self) -> None:
        with self._connect() as conn:
            conn.execute(_SCHEMA_APPLICATIONS)
            conn.execute(_SCHEMA_FOLLOW_UPS)

    # -- applications --

    def add_application(self, **kwargs: Any) -> int:
        """Insert a new application row. Returns the new row id."""
        cols = [k for k in kwargs if kwargs[k] is not None]
        placeholders = ", ".join("?" for _ in cols)
        col_names = ", ".join(cols)
        vals = [kwargs[c] for c in cols]
        with self._connect() as conn:
            cur = conn.execute(
                f"INSERT INTO applications ({col_names}) VALUES ({placeholders})", vals
            )
            return cur.lastrowid  # type: ignore[return-value]

    def update_application(self, app_id: int, **kwargs: Any) -> bool:
        """Update fields on an existing application. Returns True if a row was modified."""
        sets = ", ".join(f"{k} = ?" for k in kwargs)
        vals = list(kwargs.values()) + [app_id]
        with self._connect() as conn:
            cur = conn.execute(
                f"UPDATE applications SET {sets} WHERE id = ? AND archived = 0", vals
            )
            return cur.rowcount > 0

    def get_application(self, app_id: int) -> dict | None:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM applications WHERE id = ? AND archived = 0", (app_id,)
            ).fetchone()
            return dict(row) if row else None

    def find_by_company(self, company: str) -> list[dict]:
        """Fuzzy-find applications by company name (non-archived)."""
        norm = _normalize_company(company)
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM applications WHERE archived = 0 ORDER BY id DESC"
            ).fetchall()
        return [dict(r) for r in rows if _normalize_company(r["company"]) == norm]

    def search(self, term: str) -> list[dict]:
        """Search company and position columns."""
        pattern = f"%{term}%"
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM applications WHERE archived = 0 "
                "AND (company LIKE ? OR position LIKE ?) "
                "ORDER BY id DESC",
                (pattern, pattern),
            ).fetchall()
        return [dict(r) for r in rows]

    def list_active(self) -> list[dict]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM applications WHERE archived = 0 "
                "AND status NOT IN ('rejected', 'withdrawn') "
                "ORDER BY date_updated DESC, id DESC"
            ).fetchall()
        return [dict(r) for r in rows]

    def list_by_status(self, status: str) -> list[dict]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM applications WHERE archived = 0 AND status = ? "
                "ORDER BY date_updated DESC", (status,)
            ).fetchall()
        return [dict(r) for r in rows]

    def list_by_source(self, source: str) -> list[dict]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM applications WHERE archived = 0 "
                "AND LOWER(source) = LOWER(?) ORDER BY id DESC", (source,)
            ).fetchall()
        return [dict(r) for r in rows]

    def list_all(self, include_archived: bool = False) -> list[dict]:
        clause = "" if include_archived else "WHERE archived = 0"
        with self._connect() as conn:
            rows = conn.execute(
                f"SELECT * FROM applications {clause} ORDER BY id DESC"
            ).fetchall()
        return [dict(r) for r in rows]

    def count_since(self, since_iso: str) -> int:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT COUNT(*) FROM applications WHERE archived = 0 "
                "AND date_found >= ?", (since_iso,)
            ).fetchone()
            return row[0] if row else 0

    def stats(self) -> dict[str, int]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT status, COUNT(*) FROM applications WHERE archived = 0 "
                "GROUP BY status"
            ).fetchall()
        return {r[0]: r[1] for r in rows}

    def archive(self, app_id: int) -> bool:
        with self._connect() as conn:
            cur = conn.execute(
                "UPDATE applications SET archived = 1 WHERE id = ?", (app_id,)
            )
            return cur.rowcount > 0

    # -- follow-ups --

    def add_follow_up(self, application_id: int, due_date: str,
                      note: str = "") -> int:
        with self._connect() as conn:
            cur = conn.execute(
                "INSERT INTO follow_ups (application_id, due_date, note) "
                "VALUES (?, ?, ?)",
                (application_id, due_date, note),
            )
            return cur.lastrowid  # type: ignore[return-value]

    def get_pending_follow_ups(self) -> list[dict]:
        """Return incomplete follow-ups with application info, ordered by due date."""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT f.*, a.company, a.position FROM follow_ups f "
                "JOIN applications a ON f.application_id = a.id "
                "WHERE f.completed = 0 AND a.archived = 0 "
                "ORDER BY f.due_date ASC"
            ).fetchall()
        return [dict(r) for r in rows]

    def get_overdue_follow_ups(self) -> list[dict]:
        today = _today_iso()
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT f.*, a.company, a.position FROM follow_ups f "
                "JOIN applications a ON f.application_id = a.id "
                "WHERE f.completed = 0 AND f.due_date <= ? AND a.archived = 0 "
                "ORDER BY f.due_date ASC",
                (today,),
            ).fetchall()
        return [dict(r) for r in rows]

    def complete_follow_up(self, follow_up_id: int) -> bool:
        with self._connect() as conn:
            cur = conn.execute(
                "UPDATE follow_ups SET completed = 1 WHERE id = ?", (follow_up_id,)
            )
            return cur.rowcount > 0


# ── LLM extraction prompt ────────────────────────────────────────────────────

_EXTRACT_SYSTEM = """\
You are a structured data extractor for a job application tracker.
Given a user command, extract the requested fields as a JSON object.
Return ONLY valid JSON, no markdown, no explanation.
If a field cannot be determined, omit it from the JSON."""

_EXTRACT_ADD_PROMPT = """\
Extract job application details from this command.

Fields to extract:
- company (string, REQUIRED)
- position (string, REQUIRED)
- source (string: linkedin, indeed, recruiter, direct, referral, etc.)
- location (string)
- method (string: online, email, recruiter, referral)
- salary_range (string)
- already_applied (boolean: true if user says they already applied)

Command: {command}"""

_EXTRACT_UPDATE_PROMPT = """\
Extract the company name and new status from this command.

Valid statuses: new, applied, interviewing, offered, rejected, withdrawn

Fields:
- company (string, REQUIRED)
- status (string from list above, REQUIRED)

Command: {command}"""

_EXTRACT_FOLLOW_UP_PROMPT = """\
Extract follow-up details from this command.

Fields:
- company (string, REQUIRED)
- days (integer: number of days from now for the follow-up, default 3)
- note (string: what the follow-up is about)

Command: {command}"""


# ── Talent ────────────────────────────────────────────────────────────────────

class JobTrackerTalent(BaseTalent):
    """Track job applications, follow-ups, and export reports."""

    name = "job_tracker"
    description = (
        "Track job applications, update statuses, manage follow-ups, "
        "view statistics, and export reports for job searches"
    )
    keywords = [
        "job", "application", "applied", "job tracker", "job search",
        "follow up", "interview", "applications",
    ]
    examples = [
        "add a job application at Netflix for VP of Engineering",
        "I applied to the Microsoft Azure CISO role",
        "update the Netflix application to interviewing",
        "show my active applications",
        "what jobs need follow up",
        "how many applications this week",
        "export my job tracker",
    ]
    priority = 60

    def __init__(self) -> None:
        super().__init__()
        self._db: _DB | None = None

    def initialize(self, config: dict) -> None:
        """Open (or create) the job tracker database."""
        db_path = self.talent_config.get(
            "db_path",
            os.path.join(_data_dir(), "job_tracker.db"),
        )
        try:
            self._db = _DB(db_path)
            log.info(f"[JobTracker] Database ready at {db_path}")
        except Exception as e:
            log.error(f"[JobTracker] Failed to open database: {e}")
            self._db = None

    @property
    def routing_available(self) -> bool:
        return self._db is not None

    # ── Config schema ─────────────────────────────────────────────────────────

    def get_config_schema(self) -> dict:
        return {
            "fields": [
                {
                    "key": "db_path",
                    "label": "Database Path",
                    "type": "string",
                    "default": os.path.join(_data_dir(), "job_tracker.db"),
                },
            ]
        }

    # ── Routing ───────────────────────────────────────────────────────────────

    def can_handle(self, command: str) -> bool:
        return self.keyword_match(command)

    # ── Main dispatch ─────────────────────────────────────────────────────────

    def execute(self, command: str, context: dict) -> dict:
        """Route the user command to the appropriate handler."""
        cmd = command.lower()

        # Determine intent from keywords in the command
        if "import" in cmd and ("spreadsheet" in cmd or "xlsx" in cmd or "excel" in cmd):
            return self._handle_import(command, context)
        if self._is_export(cmd):
            return self._handle_export(context)
        if self._is_follow_up(cmd):
            if self._is_query(cmd):
                return self._handle_follow_up_list()
            return self._handle_follow_up_add(command, context)
        if self._is_stats(cmd):
            return self._handle_stats(cmd)
        if self._is_update(cmd):
            return self._handle_update(command, context)
        if self._is_add(cmd):
            return self._handle_add(command, context)
        if self._is_search(cmd):
            return self._handle_search(command, context)
        if self._is_list(cmd):
            return self._handle_list(cmd)

        # Fallback: try to figure out intent via LLM
        intent = self._extract_arg(
            context["llm"], command, "intent",
            options=["add", "update", "list", "stats", "export",
                     "follow_up", "search"],
        )
        handlers = {
            "add": lambda: self._handle_add(command, context),
            "update": lambda: self._handle_update(command, context),
            "list": lambda: self._handle_list(cmd),
            "stats": lambda: self._handle_stats(cmd),
            "export": lambda: self._handle_export(context),
            "follow_up": lambda: self._handle_follow_up_add(command, context),
            "search": lambda: self._handle_search(command, context),
        }
        if intent and intent in handlers:
            return handlers[intent]()

        # Ultimate fallback: show active applications
        return self._handle_list(cmd)

    # ── Intent detection helpers ──────────────────────────────────────────────

    @staticmethod
    def _is_add(cmd: str) -> bool:
        return bool(re.search(
            r'\b(add|new|applied to|i applied|applied for|submit)\b', cmd
        ))

    @staticmethod
    def _is_update(cmd: str) -> bool:
        return bool(re.search(
            r'\b(update|mark|change status|move to|set status|got an offer|got rejected)\b',
            cmd,
        ))

    @staticmethod
    def _is_list(cmd: str) -> bool:
        return bool(re.search(
            r'\b(show|list|active|display|view|what are|my applications)\b', cmd
        ))

    @staticmethod
    def _is_stats(cmd: str) -> bool:
        return bool(re.search(
            r'\b(how many|count|stats|statistics|summary|this week|this month)\b', cmd
        ))

    @staticmethod
    def _is_export(cmd: str) -> bool:
        return bool(re.search(
            r'\b(export|spreadsheet|xlsx|generate report|download)\b', cmd
        ))

    @staticmethod
    def _is_follow_up(cmd: str) -> bool:
        return bool(re.search(r'\bfollow[ -]?up\b', cmd))

    @staticmethod
    def _is_query(cmd: str) -> bool:
        return bool(re.search(
            r'\b(show|list|what|need|overdue|pending|check)\b', cmd
        ))

    @staticmethod
    def _is_search(cmd: str) -> bool:
        return bool(re.search(r'\b(find|search|look for|lookup)\b', cmd))

    # ── Add ───────────────────────────────────────────────────────────────────

    def _handle_add(self, command: str, context: dict) -> dict:
        """Add a new job application."""
        llm = context["llm"]
        data = self._extract_json(llm, _EXTRACT_ADD_PROMPT.format(command=command))
        if not data or not data.get("company") or not data.get("position"):
            return self._fail("I need at least a company and position. "
                              "Try: 'add a job at Netflix for Senior Engineer'")

        today = _today_iso()
        already_applied = data.get("already_applied", False)

        row = {
            "company": data["company"],
            "position": data["position"],
            "location": data.get("location", ""),
            "source": data.get("source", ""),
            "method": data.get("method", ""),
            "salary_range": data.get("salary_range", ""),
            "status": "applied" if already_applied else "new",
            "date_found": today,
            "date_applied": today if already_applied else "",
            "date_updated": today,
        }

        app_id = self._db.add_application(**row)
        status_note = " (marked as applied)" if already_applied else ""

        return {
            "success": True,
            "response": (
                f"Added application #{app_id}: **{data['position']}** at "
                f"**{data['company']}**{status_note}."
            ),
            "actions_taken": [{"action": "job_add", "id": app_id,
                               "company": data["company"]}],
            "spoken": False,
        }

    # ── Update status ─────────────────────────────────────────────────────────

    def _handle_update(self, command: str, context: dict) -> dict:
        """Update the status of an existing application."""
        llm = context["llm"]
        data = self._extract_json(llm, _EXTRACT_UPDATE_PROMPT.format(command=command))
        if not data or not data.get("company") or not data.get("status"):
            return self._fail("I need a company name and new status. "
                              "Try: 'update Netflix to interviewing'")

        new_status = data["status"].lower().strip()
        if new_status not in VALID_STATUSES:
            return self._fail(
                f"'{new_status}' is not a valid status. "
                f"Options: {', '.join(VALID_STATUSES)}"
            )

        matches = self._db.find_by_company(data["company"])
        if not matches:
            # Try a broader search
            matches = self._db.search(data["company"])
        if not matches:
            return self._fail(
                f"No application found matching '{data['company']}'. "
                "Use 'show my applications' to see what's tracked."
            )

        app = matches[0]  # Most recent match

        # Validate transition
        current = app["status"]
        allowed = _STATUS_TRANSITIONS.get(current, set())
        if new_status != current and new_status not in allowed:
            return self._fail(
                f"Cannot move from '{current}' to '{new_status}'. "
                f"Allowed transitions from '{current}': "
                f"{', '.join(sorted(allowed)) or 'none (terminal status)'}."
            )

        update_fields: dict[str, Any] = {
            "status": new_status,
            "date_updated": _today_iso(),
        }
        if new_status == "applied" and not app["date_applied"]:
            update_fields["date_applied"] = _today_iso()

        self._db.update_application(app["id"], **update_fields)

        return {
            "success": True,
            "response": (
                f"Updated **{app['company']}** ({app['position']}) "
                f"from *{current}* to *{new_status}*."
            ),
            "actions_taken": [{"action": "job_update", "id": app["id"],
                               "from": current, "to": new_status}],
            "spoken": False,
        }

    # ── List ──────────────────────────────────────────────────────────────────

    def _handle_list(self, cmd: str) -> dict:
        """List applications, with optional filtering."""
        # Check for source filter: "from linkedin"
        source_match = re.search(r'\bfrom\s+(\w+)', cmd)
        if source_match:
            source = source_match.group(1)
            apps = self._db.list_by_source(source)
            title = f"Applications from {source}"
        # Check for status filter
        elif any(s in cmd for s in VALID_STATUSES):
            status = next(s for s in VALID_STATUSES if s in cmd)
            apps = self._db.list_by_status(status)
            title = f"Applications with status '{status}'"
        else:
            apps = self._db.list_active()
            title = "Active applications"

        if not apps:
            return {
                "success": True,
                "response": f"No applications found ({title.lower()}).",
                "actions_taken": [{"action": "job_list"}],
                "spoken": False,
            }

        lines = [f"**{title}** ({len(apps)}):\n"]
        for app in apps:
            status_icon = {
                "new": "o", "applied": ">", "interviewing": "?",
                "offered": "$", "rejected": "x", "withdrawn": "-",
            }.get(app["status"], " ")
            date_str = _friendly_date(app.get("date_updated") or app.get("date_found"))
            lines.append(
                f"[{status_icon}] #{app['id']} **{app['company']}** -- "
                f"{app['position']} ({app['status']}) {date_str}"
            )

        return {
            "success": True,
            "response": "\n".join(lines),
            "actions_taken": [{"action": "job_list", "count": len(apps)}],
            "spoken": False,
        }

    # ── Stats ─────────────────────────────────────────────────────────────────

    def _handle_stats(self, cmd: str) -> dict:
        """Show application statistics."""
        status_counts = self._db.stats()
        total = sum(status_counts.values())

        today = date.today()
        week_start = (today - timedelta(days=today.weekday())).isoformat()
        month_start = today.replace(day=1).isoformat()

        this_week = self._db.count_since(week_start)
        this_month = self._db.count_since(month_start)

        lines = [f"**Job Search Stats** (total: {total})\n"]
        for status in VALID_STATUSES:
            count = status_counts.get(status, 0)
            if count > 0:
                lines.append(f"  {status}: {count}")
        lines.append(f"\nThis week: {this_week}")
        lines.append(f"This month: {this_month}")

        # Overdue follow-ups
        overdue = self._db.get_overdue_follow_ups()
        if overdue:
            lines.append(f"\n{len(overdue)} overdue follow-up(s)!")

        return {
            "success": True,
            "response": "\n".join(lines),
            "actions_taken": [{"action": "job_stats"}],
            "spoken": False,
        }

    # ── Search ────────────────────────────────────────────────────────────────

    def _handle_search(self, command: str, context: dict) -> dict:
        """Search applications by company or position text."""
        llm = context["llm"]
        term = self._extract_arg(llm, command, "search term") or ""
        if not term:
            # Try stripping common prefixes
            for prefix in ("find", "search", "look for", "lookup",
                           "find applications for", "search for"):
                if command.lower().startswith(prefix):
                    term = command[len(prefix):].strip()
                    break
        if not term:
            return self._fail("What should I search for? "
                              "Try: 'find applications for engineer'")

        apps = self._db.search(term)
        if not apps:
            return {
                "success": True,
                "response": f"No applications matching '{term}'.",
                "actions_taken": [{"action": "job_search", "term": term}],
                "spoken": False,
            }

        lines = [f"**Search results for '{term}'** ({len(apps)}):\n"]
        for app in apps:
            lines.append(
                f"  #{app['id']} **{app['company']}** -- {app['position']} "
                f"({app['status']})"
            )

        return {
            "success": True,
            "response": "\n".join(lines),
            "actions_taken": [{"action": "job_search", "term": term,
                               "count": len(apps)}],
            "spoken": False,
        }

    # ── Follow-ups ────────────────────────────────────────────────────────────

    def _handle_follow_up_add(self, command: str, context: dict) -> dict:
        """Add a follow-up reminder for an application."""
        llm = context["llm"]
        data = self._extract_json(
            llm, _EXTRACT_FOLLOW_UP_PROMPT.format(command=command)
        )
        if not data or not data.get("company"):
            return self._fail("I need a company name. "
                              "Try: 'add a follow up for Netflix in 5 days'")

        matches = self._db.find_by_company(data["company"])
        if not matches:
            matches = self._db.search(data["company"])
        if not matches:
            return self._fail(f"No application found for '{data['company']}'.")

        app = matches[0]
        days = int(data.get("days", 3))
        due = (date.today() + timedelta(days=days)).isoformat()
        note = data.get("note", "")

        fu_id = self._db.add_follow_up(app["id"], due, note)

        return {
            "success": True,
            "response": (
                f"Follow-up #{fu_id} set for **{app['company']}** "
                f"on {_friendly_date(due)} ({due})."
                + (f"\nNote: {note}" if note else "")
            ),
            "actions_taken": [{"action": "job_follow_up_add", "id": fu_id,
                               "application_id": app["id"]}],
            "spoken": False,
        }

    def _handle_follow_up_list(self) -> dict:
        """Show pending and overdue follow-ups."""
        overdue = self._db.get_overdue_follow_ups()
        pending = self._db.get_pending_follow_ups()

        if not pending:
            return {
                "success": True,
                "response": "No pending follow-ups.",
                "actions_taken": [{"action": "job_follow_up_list"}],
                "spoken": False,
            }

        lines = [f"**Pending follow-ups** ({len(pending)}):\n"]
        for fu in pending:
            is_overdue = fu["due_date"] <= _today_iso()
            marker = " [OVERDUE]" if is_overdue else ""
            lines.append(
                f"  #{fu['id']} **{fu['company']}** -- {fu['position']} "
                f"due {_friendly_date(fu['due_date'])}{marker}"
                + (f"  ({fu['note']})" if fu["note"] else "")
            )

        return {
            "success": True,
            "response": "\n".join(lines),
            "actions_taken": [{"action": "job_follow_up_list",
                               "count": len(pending),
                               "overdue": len(overdue)}],
            "spoken": False,
        }

    # ── Export ────────────────────────────────────────────────────────────────

    def _handle_import(self, command: str, context: dict) -> dict:
        """Import applications from an existing XLSX file.

        Matches the 2026-JobSearch.xlsx format:
          Sheet 1 columns: Company, title, date applied, found on,
                           applied through, Status, Notes
        """
        try:
            import openpyxl
        except ImportError:
            return self._fail("openpyxl is not installed. Run: pip install openpyxl")

        # Try to find the file path from the command or use default
        llm = context.get("llm")
        file_path = None

        # Check for common locations
        candidates = [
            os.path.expanduser("~/OneDrive/Documents/2026-JobSearch.xlsx"),
            os.path.expanduser("~/Documents/2026-JobSearch.xlsx"),
            os.path.expanduser("~/Desktop/2026-JobSearch.xlsx"),
        ]
        for c in candidates:
            if os.path.exists(c):
                file_path = c
                break

        if not file_path:
            return self._fail(
                "Could not find 2026-JobSearch.xlsx. "
                "Checked ~/OneDrive/Documents, ~/Documents, ~/Desktop."
            )

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb["Jobs Applied For"] if "Jobs Applied For" in wb.sheetnames else wb.active

            imported = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue

                company = str(row[0]).strip() if row[0] else ""
                position = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if not company or not position:
                    skipped += 1
                    continue

                # Parse date
                date_applied = ""
                if len(row) > 2 and row[2]:
                    if hasattr(row[2], 'isoformat'):
                        date_applied = row[2].strftime("%Y-%m-%d")
                    else:
                        date_applied = str(row[2])[:10]

                source = str(row[3]).strip().lower() if len(row) > 3 and row[3] else ""
                method = str(row[4]).strip().lower() if len(row) > 4 and row[4] else "direct"

                # Status mapping from spreadsheet to internal
                raw_status = str(row[5]).strip().lower() if len(row) > 5 and row[5] else ""
                status = "applied"  # default
                if "reject" in raw_status:
                    status = "rejected"
                elif "interview" in raw_status or "screen" in raw_status:
                    status = "interviewing"
                elif "offer" in raw_status:
                    status = "offered"
                elif "withdraw" in raw_status:
                    status = "withdrawn"

                notes = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                # The Status column often contains interview notes
                if raw_status and status == "interviewing":
                    notes = raw_status + (f"; {notes}" if notes else "")

                # Check for duplicate
                existing = self._db.search(company)
                dupe = any(
                    _normalize_company(e["company"]) == _normalize_company(company)
                    and e["position"].lower() == position.lower()
                    for e in existing
                )
                if dupe:
                    skipped += 1
                    continue

                self._db.add(
                    company=company,
                    position=position,
                    source=source,
                    method=method,
                    status=status,
                    date_applied=date_applied,
                    date_found=date_applied,
                    notes=notes,
                )
                imported += 1

            wb.close()
            return {
                "success": True,
                "response": (
                    f"Imported {imported} application(s) from {os.path.basename(file_path)}. "
                    f"{skipped} skipped (duplicates or incomplete rows)."
                ),
                "actions_taken": [{"action": "job_import", "path": file_path,
                                   "imported": imported, "skipped": skipped}],
                "spoken": False,
            }
        except Exception as e:
            return self._fail(f"Import failed: {e}")

    def _handle_export(self, context: dict) -> dict:
        """Export applications to XLSX matching the unemployment reporting format.

        Matches the user's existing 2026-JobSearch.xlsx layout:
          Sheet 1 "Jobs Applied For": Company, title, date applied, found on,
                                       applied through, Status, Notes
          Sheet 2 "Recruitment Sites Submitted": Firm, Date, Contact?
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return self._fail(
                "openpyxl is not installed. Run: pip install openpyxl"
            )

        apps = self._db.list_all(include_archived=False)
        if not apps:
            return self._fail("No applications to export.")

        wb = openpyxl.Workbook()

        # ── Sheet 1: Jobs Applied For ─────────────────────────────────
        ws = wb.active
        ws.title = "Jobs Applied For"

        headers = ["Company", "title", "date applied", "found on",
                    "applied through", "Status", "Notes"]
        widths = [22, 40, 14, 14, 16, 20, 40]

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                   fill_type="solid")
        thin_border = Border(bottom=Side(style="thin"))

        for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, app in enumerate(apps, 2):
            ws.cell(row=row_idx, column=1, value=app.get("company", ""))
            ws.cell(row=row_idx, column=2, value=app.get("position", ""))
            # Date as datetime for Excel formatting
            date_val = app.get("date_applied") or app.get("date_found", "")
            if date_val:
                try:
                    date_val = datetime.fromisoformat(date_val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_idx, column=3, value=date_val)
            ws.cell(row=row_idx, column=4, value=app.get("source", ""))
            ws.cell(row=row_idx, column=5, value=app.get("method", "direct"))
            # Status — map internal values to display
            status = app.get("status", "")
            status_display = {
                "new": "", "applied": "", "interviewing": "",
                "offered": "Offer", "rejected": "Rejection",
                "withdrawn": "Withdrawn",
            }.get(status, status)
            # Include contact/interview notes in status if present
            if status == "interviewing" and app.get("notes"):
                status_display = app["notes"]
            elif status == "rejected":
                status_display = "Rejection"
            ws.cell(row=row_idx, column=6, value=status_display)
            ws.cell(row=row_idx, column=7, value=app.get("notes", ""))

        # Freeze header row
        ws.freeze_panes = "A2"

        # ── Sheet 2: Recruitment Sites Submitted ──────────────────────
        ws2 = wb.create_sheet("Recruitment Sites Submitted")
        rec_headers = ["Firm", "Date", "Contact?"]
        rec_widths = [25, 14, 14]

        for col_idx, (header, width) in enumerate(zip(rec_headers, rec_widths), 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            ws2.column_dimensions[get_column_letter(col_idx)].width = width

        # Populate from applications sourced via recruiters
        recruiter_apps = [a for a in apps
                          if (a.get("source", "").lower() in
                              ("recruiter", "staffing", "headhunter")
                              or a.get("method", "").lower() == "recruiter")]
        seen_firms = set()
        rec_row = 2
        for app in recruiter_apps:
            firm = app.get("company", "")
            if firm.lower() in seen_firms:
                continue
            seen_firms.add(firm.lower())
            ws2.cell(row=rec_row, column=1, value=firm)
            date_val = app.get("date_applied") or app.get("date_found", "")
            if date_val:
                try:
                    date_val = datetime.fromisoformat(date_val)
                except (ValueError, TypeError):
                    pass
            ws2.cell(row=rec_row, column=2, value=date_val)
            ws2.cell(row=rec_row, column=3,
                     value=app.get("contact_name", ""))
            rec_row += 1

        ws2.freeze_panes = "A2"

        # Save
        filename = f"job_tracker_export_{_today_iso()}.xlsx"
        filepath = os.path.join(_data_dir(), filename)
        wb.save(filepath)

        return {
            "success": True,
            "response": (
                f"Exported {len(apps)} application(s) to:\n`{filepath}`\n"
                f"Matches your unemployment reporting format (2 sheets)."
            ),
            "actions_taken": [{"action": "job_export", "path": filepath,
                               "count": len(apps)}],
            "spoken": False,
        }

    # ── Cowork integration ────────────────────────────────────────────────────

    def add_from_cowork(self, data: dict) -> int:
        """Add an application from Cowork bridge data.

        Args:
            data: Dict with keys matching the applications schema columns.

        Returns:
            The new application row id.
        """
        if not self._db:
            raise RuntimeError("JobTracker database not initialized")

        # Ensure required fields
        if not data.get("company") or not data.get("position"):
            raise ValueError("company and position are required")

        # Set defaults
        data.setdefault("date_found", _today_iso())
        data.setdefault("date_updated", _today_iso())
        data.setdefault("status", "new")

        # Filter to only valid columns
        valid_cols = {
            "company", "position", "location", "source", "status",
            "date_found", "date_applied", "date_updated", "contact_name",
            "contact_email", "method", "salary_range", "notes", "job_url",
            "resume_version", "cover_letter", "cowork_task_id", "fit_score",
        }
        filtered = {k: v for k, v in data.items() if k in valid_cols}
        return self._db.add_application(**filtered)

    def get_active_applications(self) -> list[dict]:
        """Return all non-archived, non-terminal applications.

        Used by Cowork to check for duplicates before adding.
        """
        if not self._db:
            return []
        return self._db.list_active()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _extract_json(self, llm: Any, prompt: str) -> dict | None:
        """Ask the LLM to extract structured data as JSON."""
        try:
            raw = llm.generate(
                prompt,
                system_prompt=_EXTRACT_SYSTEM,
                temperature=0.1,
                max_length=256,
            )
        except LLMError as e:
            log.error(f"[JobTracker] LLM extraction failed: {e}")
            return None

        try:
            clean = raw.strip()
            if clean.startswith("```"):
                clean = re.sub(r"^```[a-z]*\n?", "", clean)
                clean = re.sub(r"\n?```$", "", clean.strip())
            match = re.search(r'\{.*\}', clean, re.DOTALL)
            if match:
                return json.loads(match.group())
        except (json.JSONDecodeError, AttributeError) as e:
            log.error(f"[JobTracker] JSON parse failed: {e}")

        return None

    # ── Claude CLI integration ─────────────────────────────────────────────

    @staticmethod
    def _claude_generate(prompt: str, timeout: float = 60.0) -> str | None:
        """Call `claude -p` for writing tasks that need better than local LLM.

        Returns the response text, or None on failure.
        Uses the Claude CLI in pipe mode — works independently of Claude Code.
        """
        import subprocess
        try:
            result = subprocess.run(
                ["claude", "-p", prompt],
                capture_output=True, text=True, timeout=timeout,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
            log.warning(f"[JobTracker] claude -p returned code {result.returncode}")
            return None
        except FileNotFoundError:
            log.warning("[JobTracker] claude CLI not found — install with: npm i -g @anthropic-ai/claude-code")
            return None
        except subprocess.TimeoutExpired:
            log.warning("[JobTracker] claude -p timed out")
            return None
        except Exception as e:
            log.error(f"[JobTracker] claude -p error: {e}")
            return None

    def evaluate_fit(self, app_id: int, job_description: str,
                     resume_summary: str) -> dict:
        """Ask Claude to evaluate how well a job matches the user's resume.

        Returns dict with fit_score (0-100) and analysis text.
        Updates the application's fit_score in the database.
        """
        prompt = (
            "You are evaluating job fit. Given a resume summary and job description, "
            "return a JSON object with:\n"
            '  {"fit_score": <0-100>, "analysis": "<2-3 sentences on fit>",'
            ' "strengths": ["..."], "gaps": ["..."]}\n\n'
            f"RESUME SUMMARY:\n{resume_summary}\n\n"
            f"JOB DESCRIPTION:\n{job_description[:3000]}\n\n"
            "Return ONLY JSON."
        )
        raw = self._claude_generate(prompt)
        if not raw:
            return {"fit_score": 0, "analysis": "Claude CLI unavailable."}

        try:
            clean = re.sub(r'^```[a-z]*\n?', '', raw.strip())
            clean = re.sub(r'\n?```$', '', clean.strip())
            m = re.search(r'\{.*\}', clean, re.DOTALL)
            if m:
                result = json.loads(m.group())
                score = int(result.get("fit_score", 0))
                # Update DB
                if self._db:
                    self._db.update(app_id, fit_score=score)
                return result
        except Exception as e:
            log.error(f"[JobTracker] Fit evaluation parse error: {e}")

        return {"fit_score": 0, "analysis": "Could not parse evaluation."}

    def draft_follow_up(self, app_id: int) -> str | None:
        """Ask Claude to draft a follow-up email for an application."""
        if not self._db:
            return None
        app = self._db.get(app_id)
        if not app:
            return None

        days_since = ""
        if app["date_applied"]:
            try:
                applied = date.fromisoformat(app["date_applied"])
                days_since = f" ({(date.today() - applied).days} days since application)"
            except ValueError:
                pass

        prompt = (
            f"Draft a brief, professional follow-up email for a job application.\n\n"
            f"Company: {app['company']}\n"
            f"Position: {app['position']}\n"
            f"Applied: {app['date_applied'] or 'recently'}{days_since}\n"
            f"Contact: {app['contact_name'] or 'Hiring Manager'}\n\n"
            "Keep it concise (3-4 sentences), professional, and enthusiastic. "
            "Do not include a subject line. Do not include a signature — the user will add their own."
        )
        return self._claude_generate(prompt)

    # ── Cowork bridge methods ────────────────────────────────────────────

    def send_to_cowork(self, task_type: str, payload: dict) -> str | None:
        """Write a task JSON for Cowork to pick up via the bridge.

        Returns the task_id, or None on failure.
        """
        import uuid
        from pathlib import Path

        bridge_tasks = Path.home() / "OneDrive" / "Documents" / "cowork_bridge" / "tasks"
        bridge_tasks.mkdir(parents=True, exist_ok=True)

        task_id = f"job_{task_type}_{uuid.uuid4().hex[:8]}"
        task = {
            "task_id": task_id,
            "task_type": task_type,
            "created": datetime.now().isoformat(),
            "payload": payload,
        }
        task_path = bridge_tasks / f"{task_id}.json"
        with open(task_path, "w", encoding="utf-8") as f:
            json.dump(task, f, indent=2)
        log.info(f"[JobTracker] Cowork task written: {task_path.name}")
        return task_id

    def request_resume_tailoring(self, app_id: int,
                                  job_description: str) -> str | None:
        """Send a resume tailoring request to Cowork for a specific application.

        Cowork will use the user's master resume + JD to produce tailored materials.
        Returns the Cowork task_id.
        """
        if not self._db:
            return None
        app = self._db.get(app_id)
        if not app:
            return None

        task_id = self.send_to_cowork("resume_tailor", {
            "application_id": app_id,
            "company": app["company"],
            "position": app["position"],
            "job_url": app.get("job_url", ""),
            "job_description": job_description[:5000],
            "resume_path": "~/OneDrive/Documents/resume_master.md",
            "instructions": (
                "Read the user's master resume from the resume_path. "
                "Tailor it for this position. Also draft a cover letter. "
                "Follow the writing style rules in ~/.claude/CLAUDE.md. "
                "Return both as separate sections."
            ),
        })
        if task_id and self._db:
            self._db.update(app_id, cowork_task_id=task_id)
        return task_id

    @staticmethod
    def _fail(message: str) -> dict:
        """Return a standard failure result."""
        return {
            "success": False,
            "response": message,
            "actions_taken": [],
            "spoken": False,
        }
